import xml.etree.ElementTree as etree
import matplotlib.pyplot as plt
import numpy as np
from numpy import exp
from lmfit import Model
import pandas as pd
from openpyxl import load_workbook
import statsmodels.api as sm
from openpyxl.utils import get_column_letter

xml = etree.parse('HY202103_D08_(0,2)_LION1_DCM_LMZC.xml')
root = xml.getroot()

def poly(x, y, degree):
    coeffs = np.polyfit(x, y, degree)
    # r-squared
    p = np.poly1d(coeffs)
    yhat = p(x)
    ybar = np.sum(y) / len(y)           # or sum(y)/len(y)
    ssreg = np.sum((yhat - ybar) ** 2)  # or sum([ (yihat - ybar)**2 for yihat in yhat])
    sstot = np.sum((y - ybar) ** 2)     # or sum([ (yi - ybar)**2 for yi in y])
    results = ssreg / sstot
    return results

# I-V graph
V = []
for v in root.iter('Voltage'):
    V.extend(list(map(float, v.text.split(','))))
I = []
for i in root.iter('Current'):
    I.extend(list(map(float, i.text.split(','))))
    I = list(map(abs, I))

x = np.array(V[:])
y = np.array(I[:])
fit1 = np.polyfit(x, y, 12)
fit1 = np.poly1d(fit1)

def IV_fit(X, Is, Vt):
    return (Is * (exp(X/Vt) - 1) + fit1(X))

model = Model(IV_fit)
result = model.fit(y, X=x, Is=10**-15, Vt=0.026)

# R-squared
def IVR(y):
    yhat = result.best_fit
    ybar = np.sum(y)/len(y)
    sse = np.sum((yhat - ybar) ** 2)
    sst = np.sum((y - ybar) ** 2)
    return sse/sst

# R-squared code
# r2 = r2_score(I, np.polyval(fit1, V))
# print(r2)

plt.subplot(2, 3, 4)
plt.plot(V, I, 'b.', label='data', markersize=8)
plt.plot(x, result.best_fit, label='best_fit')
plt.plot(x, result.best_fit, 'r-', label='R-squared ={}'.format(IVR(y)))
plt.legend(loc = 'best', fontsize=6)
plt.yscale('log')
plt.title('I-V analysis', fontsize=12, fontweight='bold')
plt.xlabel('Voltage[V]', fontsize=8)
plt.ylabel('Current[A]', fontsize=8)

# wavelength-transmission graph
counts = []
for i in root.iter('WavelengthSweep'):
    site_values = list(i.attrib.values())
    counts.append(site_values)

plt.subplot(2, 3, 1)
L = []
IL = []
DCBias = []
for i in root.iter('WavelengthSweep'):
    L.append(list(map(float, i[0].text.split(','))))
    IL.append(list(map(float, i[1].text.split(','))))
    site_key = list(i.attrib.keys())
    site_values = list(i.attrib.values())
    DCBias.append(site_key[1]+'='+site_values[1])
DCBias[-1] = 'Reference'
for r in range(len(L)):
    plt.plot(L[r], IL[r], label=DCBias[r])
plt.xlabel('Wavelength[nm]', fontsize=8)
plt.ylabel('Transmission[dB]', fontsize=8)
plt.title('Transmission spectra as measured', fontsize=12, fontweight='bold')
plt.legend(loc='best', ncol=2, fontsize=6)

# reference fitting
plt.subplot(2, 3, 2)
plt.plot(L[-1], IL[-1], 'r', label = 'reference')
fit_L = np.polyfit(L[-1], IL[-1], 2)
fit_IL = np.polyval(fit_L, L[-1])
plt.plot(L[-1], fit_IL, 'c-', label='Fit ref polynomial O3')
plt.legend(loc='best', fontsize=6)
plt.xlabel('Wavelength[nm]', fontsize=8)
plt.ylabel('Transmission[dB]', fontsize=8)
plt.title('Transmission spectra as measured\n(Fitting)', fontsize=12, fontweight='bold')
plt.ylim(-60, -5)

# reference - fitting
plt.subplot(2, 3, 3)
del IL[2][-1]
del L[2][-1]
for j in range(len(L)-1):
    IL_f = []
    for k in range(len(IL[j])):
        IL_f.append(IL[j][k]-fit_IL[k])
    plt.plot(L[j], IL_f, label=DCBias[j])
plt.xlabel('Wavelength[nm]', fontsize=8)
plt.ylabel('Transmission[dB]', fontsize=8)
plt.title('Transmission spectra as measured\n(Fitting)', fontsize=12, fontweight='bold')
plt.legend(loc='best', ncol=2, fontsize=6)
plt.tight_layout()

plt.savefig('./graph.png')

TestSiteInfo=root.find('TestSiteInfo')
dict={'Lot':[],'Wafer':[],'Mask':[],'TestSite':[],'Name':[],'Date':[],'Operator':[],'DieRow':[],'DieColumn':[],'AnalysisWavelength (nm)':[],'Rsq of Ref. spectrum (6th)':[] ,'Max transmission of Ref. spec. (dB)':[], 'Rsq of IV':[], '1 at -1V[A]':[], '1 at 1V[A]':[]}
AlignWavelength=next(root.iter('AlignWavelength'))
Modulator=next(root.iter('Modulator'))

# rsqref = poly(L[-1], IL, 6)
IVdic = {I: V for V, I in zip(result.best_fit, V)}

initial_list = []
for i in x:
    x_value = IV_fit(i, 10e-16, 1 / 0.026)
    initial_list.append(x_value)

initial = sm.add_constant(np.abs(y))
result1 = sm.OLS(initial_list, initial).fit()

Rsq = result1.rsquared

for i in range(len(L)-1):
    dict['Lot'].append(TestSiteInfo.attrib['Batch'])
    dict['Wafer'].append(TestSiteInfo.attrib['Wafer'])
    dict['Mask'].append(TestSiteInfo.attrib['Maskset'])
    dict['TestSite'].append(TestSiteInfo.attrib['TestSite'])
    dict['Name'].append(Modulator.attrib['Name'])
    dict['Date'].append(root.attrib['CreationDate'])
    dict['Operator'].append(root.attrib['Operator'])
    dict['DieRow'].append(TestSiteInfo.attrib['DieRow'])
    dict['DieColumn'].append(TestSiteInfo.attrib['DieColumn'])
    dict['AnalysisWavelength (nm)'].append(AlignWavelength.text)
    dict['Rsq of Ref. spectrum (6th)'].append(' ')
    dict['Max transmission of Ref. spec. (dB)'].append(max(IL[i] - fit_IL))
    dict['Rsq of IV'].append(Rsq)
    dict['1 at -1V[A]'].append(IVdic[-1.0])
    dict['1 at 1V[A]'].append(IVdic[1.0])
frame=pd.DataFrame(dict)
frame.to_excel('Data.xlsx',index=False)
wb=load_workbook('Data.xlsx')
ws=wb.active
for i in ['E','F','J','K','L']:
    ws.column_dimensions[i].width = 30
ws.column_dimensions['D'].width = 13
ws.column_dimensions['H'].width = 13
wb.save('./Data1.xlsx')

# plt.show()