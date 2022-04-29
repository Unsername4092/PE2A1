import xml.etree.ElementTree as elemTree
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from sklearn.metrics import r2_score
from lmfit.models import ExpressionModel
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import string
data = elemTree.parse('LION1.xml')
root = data.getroot()
# I-V graph
V = []                          # Empty list for saving voltage value
for v in root.iter('Voltage'):  # Save each voltage value in the list.
    V.extend(list(map(float, v.text.split(','))))
I = []                          # Empty list for saving current value
for i in root.iter('Current'):  # Save the each current value in the list.
    I.extend(list(map(float, i.text.split(','))))
plt.figure(1, [18, 8])          # figure 1에 그래프를 그리고 창의 크기를 18,8로 설정
plt.subplot(2, 3, 1)                               # Set the size and number of the graph window.
plt.title('I-V analysis', fontsize=18, fontweight='bold')   # Graph title
plt.xlabel('Voltage (V)', fontsize=13)                # title of x-axis
plt.ylabel('Current (A)', fontsize=13)                # title of y-axis
plt.yscale('log')                                     # Determine the y-scale
#plt.scatter(V, np.abs(I))           # Plot the Voltage and Current values.
# Regression
p = np.polyfit(V, np.abs(I), 12)       # 다항식계수 구하기
y = np.polyval(p, V)                  # fitting 한 값 데이터
r2 = r2_score(np.abs(I), y)             # 결정계수
print(r2)
#plt.plot(V, y, color='y', label='R squared {}'.format(r2))  # fitting 한값 데이터를 표시한다. label에는 결정계수 값을 넣어준다.
#  lmfit
gmod = ExpressionModel("(exp(-x / 0.026) - 1)")
result=gmod.fit(I[:6],x=V[:6])
plt.plot(V[:6], result.init_fit, '--', label='initial fit')
plt.plot(V[:6], result.best_fit, '-', label='best fit')
plt.legend(loc='best', ncol=2)
# Wavelength-Transmission graph
L = []                      # Empty list for saving wavelength
for n in root.iter('L'):    # Save the each wavelength data in the list
    L.append(list(map(float, n.text.split(','))))
del L[2][-1]
IL = []                     # Empty list for saving Measured transmission
for m in root.iter('IL'):   # Save the each transmission data in the list
    IL.append(list(map(float, m.text.split(','))))
del IL[2][-1]
wls = []                    # Empty list for saving Voltage value
for l in root.iter('WavelengthSweep'):   # Saves the voltage value to be used as a label.
    wls.append('DC = {}'.format(l.attrib['DCBias']))
wls[-1] = 'Reference'         # Set the last label as reference.
pr = np.polyfit(L[-1], IL[-1], 6)   # L IL 그래프 다항식계수구하기
yr = np.polyval(pr, L[-1])         # fitting 된 IL 데이터
r2_r = r2_score(IL[-1], yr)           # 결정계수
for sp in range(2, 5):          # sp - subplot
    plt.subplot(2, 3, sp)       # Set the size and number of the graph window.
    plt.title('Transmission spectra - as measured', fontsize=18, fontweight='bold')   # Graph title
    plt.xlabel('Wavelength (nm)', fontsize=13)              # Title of x-axis
    plt.ylabel('Measured transmission (dBm)', fontsize=13)  # Title of y-axis
    if sp == 2:                        # 두번째 subplot 에는 reference 와 fitting 값의 그래프를 그린다.
        plt.plot(L[-1], IL[-1], color='C6')    # 기존 reference의 값이다.
        plt.plot(L[-1], yr, color='black', linestyle='dashed', label='R squared = {}'.format(r2_r)) # fitting된 그래프이다.
    if sp == 3:                        # 세번째 subplot 에는 기존 그래프에 fitting 된 reference 값을 빼준다.
        for r in range(len(L)-1):      # reference 가 빠졌기 때문에 반복횟수는 len(L)-1이된다.
            plt.plot(L[r], IL[r]-yr, label=wls[r])   # 각 원소들의 차는 List - List 로 표현할 수 있다.
    if sp == 4:                        # 네번째 subplot 에는 raw data 그래프를 그린다.
        for r in range(len(L)):
            plt.plot(L[r], IL[r], label=wls[r])
    plt.legend(loc='best', ncol=2)
plt.tight_layout()                     # 서로 title이나 축 label이 겹치치않게 한다.
#plt.savefig('./graph.png')
#plt.show()        # Show the graphs
TestSiteInfo=root.find('TestSiteInfo')
dict={'Lot':[],'Wafer':[],'Mask':[],'TestSite':[],'Name':[],'Date':[],'Operator':[],'DieColumn':[],'DieRow':[],'AnalysisWavelength (nm)':[],'Rsq of Ref. spectrum (6th)':[] ,'Max transmission of Ref. spec. (dB)':[]}
AlignWavelength=next(root.iter('AlignWavelength'))
Modulator=next(root.iter('Modulator'))
for i in range(len(L)-1):
    dict['Lot'].append(TestSiteInfo.attrib['Batch'])
    dict['Wafer'].append(TestSiteInfo.attrib['Wafer'])
    dict['Mask'].append(TestSiteInfo.attrib['Maskset'])
    dict['TestSite'].append(TestSiteInfo.attrib['TestSite'])
    dict['Name'].append(Modulator.attrib['Name'])
    dict['Date'].append(root.attrib['CreationDate'])
    dict['Operator'].append(root.attrib['Operator'])
    dict['DieColumn'].append(TestSiteInfo.attrib['DieColumn'])
    dict['DieRow'].append(TestSiteInfo.attrib['DieRow'])
    dict['AnalysisWavelength (nm)'].append(AlignWavelength.text)
    dict['Rsq of Ref. spectrum (6th)'].append(' ')
    dict['Max transmission of Ref. spec. (dB)'].append(min(IL[i]-yr))
print(dict)
frame=pd.DataFrame(dict)
frame.to_excel('Data.xlsx',index=False)
wb=load_workbook('Data.xlsx')
ws=wb.active
for i in ['E','F','J','K','L']:
    ws.column_dimensions[i].width=30
ws.column_dimensions['D'].width=13
ws.column_dimensions['H'].width=13
wb.save('Data1.xlsx')